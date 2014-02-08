# -*- coding: utf-8 -*-

# Copyright (C) 2013 Luís A. Bastião Silva and Universidade de Aveiro
#
# Authors: Luís A. Bastião Silva <bastiao@ua.pt>
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with this program.  If not, see <http://www.gnu.org/licenses/>.
#
import csv
from pprint import pprint

from django.shortcuts import render
from django.http import HttpResponse, HttpResponseRedirect
from django.core.urlresolvers import reverse
from django.db import transaction
from django.core.urlresolvers import *

from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from emif.utils import QuestionGroup, ordered_dict

from questionnaire.models import *
from questionnaire.parsers import *
from questionnaire.views import *
from questionnaire.models import *
from searchengine.search_indexes import CoreEngine
from searchengine.models import Slugs
import searchengine.search_indexes
from searchengine.search_indexes import index_answeres_from_qvalues
from searchengine.search_indexes import convert_text_to_slug
from emif.utils import *
from emif.models import *
from api.models import *

from geopy import geocoders 
from django.core.mail import send_mail, BadHeaderError

from modules.geo import *

from rest_framework.authtoken.models import Token

import json
import logging
import re
import md5
import random

import os
import os.path


def list_questions():
    print "list_questions"
    objs = Questionnaire.objects.all()
    results = {}
    for q in objs:
        results[q.id] = q.name
    print results
    return results


def index(request, template_name='index.html'):
    return render(request, template_name, {'request': request})


def about(request, template_name='about.html'):
    return render(request, template_name, {'request': request, 'breadcrumb': True})

def bootstrap_ie_compatibility(request, template_name='bootstrap_ie_compatibility.css'):
    return render(request, template_name, {'request': request, 'breadcrumb': False})

def quick_search(request, template_name='quick_search.html'):
    return render(request, template_name, {'request': request})


def results_db(request, template_name='results.html'):
    user = request.user
    su = Subject.objects.filter(user=user)
    databases = RunInfoHistory.objects.filter(subject=su)

    class Database:
        id = ''
        name = ''
        date = ''
        last_activity = ''

    class Results:
        num_results = 0
        list_results = []

    list_databases = []
    for database in databases:
        database_aux = Database()
        database_aux.id = database.runid
        database_aux.date = database.completed
        
        answers = Answer.objects.filter(runid=database.runid)
        text = clean_value(str(answers[1].answer))
        info = text[:75] + (text[75:] and '..')
        database_aux.name = info
        list_databases.append(database_aux)

    list_results = Results()
    list_results.num_results = len(list_databases)
    list_results.list_results = list_databases

    return render(request, template_name, {'request': request,
                                           'list_results': list_results})


def results_comp(request, template_name='results_comp.html'):
    list_fingerprint_to_compare = []
    print request.POST
    if request.POST:
        for k, v in request.POST.items():
            print k
            print v
            if k.startswith("chk_") and v == "on":
                arr = k.split("_")

                list_fingerprint_to_compare.append(arr[1])

    print "list_fingerprint_to_compare" + str(list_fingerprint_to_compare)

    class Results:
        num_results = 0
        list_results = []

    class DatabaseFields:
        id = ''
        name = ''
        date = ''
        fields = None

    list_qsets = []
    for db_id in list_fingerprint_to_compare:
        qsets, name, db_owners, fingerprint_ttype = createqsets(db_id)
        list_qsets.append((name, qsets))
    first_name = None
    if len(list_qsets) > 0:
        (first_name, discard) = list_qsets[0]

    print "list_qsets: " + str(list_qsets)
    return render(request, template_name, {'request': request, 'breadcrumb': True,
                                           'results': list_qsets, 'database_to_compare': first_name})


def results_fulltext(request, page=1, template_name='results.html'):
    query = ""
    in_post = True
    try:
        query = request.POST['query']
        request.session['query'] = query
    except:
        in_post = False

    if not in_post:
        query = request.session.get('query', "")

    return results_fulltext_aux(request, "text_t:" + query, page, template_name)


def results_fulltext_aux(request, query, page=1, template_name='results.html'):
    rows = 5
    if query == "":
        return render(request, "results.html", {'request': request, 'breadcrumb': True,
                                                'list_results': [], 'page_obj': None})

    class Results:
        num_results = 0
        list_results = []
        paginator = None

    c = CoreEngine()
    error_searching = False
    try:
        results = c.search_fingerprint(query, str(0))
    except: 
        error_searching = True

    questionnaires_ids = {}
    qqs = Questionnaire.objects.all()
    for q in qqs:
        questionnaires_ids[q.slug] = (q.pk, q.name)

    list_databases = []
    if error_searching or len(results) == 0 :
        query_old = request.session.get('query', "")
        return render(request, "results.html", {'request': request, 'breadcrumb': True,
                                                'list_results': [], 'page_obj': None, 'search_old': query_old})
    for r in results:
        try:
            database_aux = Database()

            if (not r.has_key('database_name_t')):
                database_aux.name = '(Unnamed)'
            else:
                database_aux.name = r['database_name_t']

            if (not r.has_key('location_t')):
                database_aux.location = ''
            else:
                database_aux.location = r['location_t']
            if (not r.has_key('institution_name_t')):
                database_aux.institution = ''
            else:
                database_aux.institution = r['institution_name_t']

            if (not r.has_key('contact_administrative_t')):
                database_aux.email_contact = ''
            else:
                database_aux.email_contact = r['contact_administrative_t']

            if (not r.has_key('number_active_patients_jan2012_t')):
                database_aux.number_patients = ''
            else:
                database_aux.number_patients = r['number_active_patients_jan2012_t']
                
            if (not r.has_key('date_last_modification_t')):
                database_aux.last_activity = ''
            else:
                database_aux.last_activity = r['date_last_modification_t']                  
                
            if (not r.has_key('upload-image_t')):
                database_aux.logo = 'nopic.gif'
            else:
                database_aux.logo = r['upload-image_t']
            database_aux.id = r['id']
            database_aux.date = convert_date(r['created_t'])
            try:
                database_aux.date_modification = convert_date(r['date_last_modification_t'])
            except KeyError:
                pass
                
            (ttype, type_name) = questionnaires_ids[r['type_t']]
            database_aux.ttype = ttype
            database_aux.type_name = type_name
            list_databases.append(database_aux)
        except:
            raise

    pp = Paginator(list_databases, rows)
    list_results = Results()
    list_results.num_results = results.hits
    list_results.list_results = pp.page(page)
    list_results.paginator = pp
    query_old = request.session.get('query', "")
    return render(request, template_name, {'request': request,
                                           'list_results': list_results, 'page_obj': pp.page(page),
                                           'search_old': query_old, 'breadcrumb': True})


def store_query(user_request, query_executed):
    print user_request.user.is_authenticated()
    print "Store Query2"
    # Verify if the query already exists in that user 
    user_aux = None
    if user_request.user.is_authenticated():

        user_aux = user_request.user

        results_tmp = QueryLog.objects.filter(query=query_executed, user=user_aux)

    else:
        results_tmp = QueryLog.objects.filter(query=query_executed, user__isnull=True)
    query = None
    
    if (results_tmp.exists()):
        # If the user exists, then update the Query 
        query = results_tmp[0]
    else:
        # Create a query 
        query = QueryLog()
        if user_request.user.is_authenticated():
            query.user = user_request.user
        else:
            query.user = None
        query.query = query_executed
        print "dmn"
    if query != None:
        if user_request.user.is_authenticated():
            query.user = user_request.user
        else:
            query.user = None
        query.query = query_executed
        query.save()


def results_diff(request, page=1, template_name='results_diff.html'):
    query = ""
    in_post = True
    try:
        query = request.POST['query']
        request.session['query'] = query
    except:
        in_post = False
        #raise
    if not in_post:
        query = request.session.get('query', "")

    if query == "":
        return render(request, "results.html", {'request': request,
                                                'list_results': [], 'page_obj': None, 'breadcrumb': True})
    store_query(request, query)
    try:
        # Store query by the user
        search_full = request.POST['search_full']
        if search_full == "search_full":
            return results_fulltext(request, page)
    except:
        return results_fulltext(request, page)


    class Results:
        num_results = 0
        list_results = []
        d1 = None
        d2 = None
        d3 = None


    class DatabaseFields:
        id = ''
        name = ''
        date = ''
        fields = None


    c = CoreEngine()
    results = c.search_fingerprint("text_t:" + query)
    
    list_databases = []

    for r in results:
        try:
            database_aux = Database()
            
            database_aux.id = r['id']
            database_aux.date = convert_date(r['created_t'])

            database_aux.name = r['database_name_t']
            database_aux.last_activity = r['date_last_modification_t']
            
            list_databases.append(database_aux)
            if (len(list_databases) == 3):
                break
        except:
            pass

    c = CoreEngine()
    list_databases_final = []
    list_results = Results()

    for db in list_databases:

        results = c.search_fingerprint('id:' + db.id)

        class Tag:
            tag = ''
            value = ''

        list_values = []
        blacklist = ['created_t', 'type_t', '_version_', 'date_last_modification_t']
        name = "Not defined"
        for result in results:
            questionnaire_slug = result['type_t']
            q_main = Questionnaire.objects.filter(slug=questionnaire_slug)[0]
            for k in result:
                if k in blacklist:
                    continue
                t = Tag()
                results = Slugs.objects.filter(slug1=k[:-2], question__questionset__questionnaire=q_main.pk)
                if len(results) > 0:
                    text = results[0].description
                else:
                    text = k
                info = text[:75] + (text[75:] and '..')

                t.tag = info

                value = clean_value(str(result[k]))
                value = value[:75] + (value[75:] and '..')
                t.value = value
                if k == "database_name_t":
                    name = t.value
                list_values.append(t)
        db.fields = list_values
        list_databases_final.append(db)

    list_results.d1 = list_databases_final[0]
    list_results.d2 = list_databases_final[1]
    list_results.d3 = list_databases_final[2]
    list_results.num_results = len(list_databases)
    return render(request, template_name, {'request': request, 'query': query,
                                           'results': list_results, 'search_old': query, 'breadcrumb': True})


def geo(request, template_name='geo.html'):
    query = None
    try:
        query = request.session['query']
        query = 'text_t:' + query
    except:
        pass
    if query == None:
        query = "*:*"
    print "query@" + query
    list_databases = get_databases_from_solr(request, query)

    list_locations = []
    _long_lats = []
    # since the geolocation is now adding the locations, we no longer need to look it up when showing,
    # we rather get it directly

    for database in list_databases:

        if database.location.find(".")!= -1:
            _loc = database.location.split(".")[0]
        else:
            _loc = database.location

        city=None
        g = geocoders.GeoNames(username='bastiao')

        if _loc!= None and g != None and len(_loc)>1:
            #try:
            #    place, (lat, lng) = g.geocode(_loc)
            #except:
            #    continue
            try:
                city = City.objects.get(name=_loc.lower())

            # if dont have this city on the db
            except City.DoesNotExist:
                print "-- Error: The city " + _loc + " doesnt exist on the database. Maybe too much requests were being made when it happened ? Trying again..."

                #obtain lat and longitude
                city = retrieve_geolocation(_loc.lower())

                if city != None:
                    print city

                    city.save()

                else:
                    print "-- Error: retrieving geolocation"
                    continue

            _long_lats.append(str(city.lat) + ", " + str(city.long))

        #print _loc

        list_locations.append(_loc)
    return render(request, template_name, {'request': request,
                                           'list_cities': list_locations, 'lats_longs': _long_lats, 'breadcrumb': True})



def statistics(request, questionnaire_id, question_set, template_name='statistics.html'):

    from emif.statistics import Statistic


    # print "QUESTIONNAIRE_ID: " + str(questionnaire_id)
    # print "QUESTION_SET: " + str(question_set)

    qs_list = QuestionSet.objects.filter(questionnaire=questionnaire_id).order_by('sortid')

    if int(question_set) == 99:
        question_set = len(qs_list) - 1
    question_set = qs_list[int(question_set)]

    questions = Question.objects.filter(questionset=question_set)

    return render(request, template_name, {'request': request, 'questionset': question_set,
                                           'breadcrumb': True, 'questions_list': questions,
                                           'questionnaire_id': questionnaire_id})


def generate_statistics_from_multiple_choice(question_slug):
    choices = Choice.objects.filter(question=q)
    total_values = calculate_total_values()
    c = CoreEngine()
    for choice in choices:
        query = "question_slug:" + "choice.value"
        results = c.search_fingerprint(query)
        number_results = len(results)


def calculate_databases_per_location():
    users = EmifProfile.objects.all()
    c = CoreEngine()
    contries = []
    for u in users:
        # Count number of DB's for each user
        query = "subject_id_t:" + u.user.id
        results = c.search_fingerprint(query)
        # Number of dbs
        number_of_dbs = len(results)
        if contries.has_key(u.contry.name):
            contries[u.contry.name] = contries[u.contry.name] + number_of_dbs
        else:
            contries[u.contry.name] = number_of_dbs


def advanced_search(request, questionnaire_id, question_set):

    return show_fingerprint_page_read_only(request, questionnaire_id, question_set, True)


def database_add(request, questionnaire_id, sortid):

    response = show_fingerprint_page_read_only(request, questionnaire_id, sortid,
                                               template_name='database_add.html')

    return response

def database_search_qs(request, questionnaire_id, sortid):

    response = render_one_questionset(request, questionnaire_id, sortid,
                                               template_name='fingerprint_search_qs.html')

    return response

def render_one_questionset(request, q_id, qs_id, errors={}, template_name='fingerprint_add_qs.html'):
    """
    Return the QuestionSet template

    Also add the javascript dependency code.
    """
    try:

        qs_list = QuestionSet.objects.filter(questionnaire=q_id, sortid=qs_id).order_by('sortid')

        if (int(qs_id) == 99):
            qs_id = len(qs_list) - 1
        question_set = qs_list[0]
        #questions = Question.objects.filter(questionset=qs_id)

        questions = question_set.questions()
        #print "Questions: " + str(questions)
        #print "QuestionSet: " + str(question_set)

        questions_list = {}
        for qset_aux in qs_list:
            #questions_aux = Question.objects.filter(questionset=qset_aux)
            questions_list[qset_aux.id] = qset_aux.questions()
            #print "here"

        qlist = []
        jsinclude = []      # js files to include
        cssinclude = []     # css files to include
        jstriggers = []
        qvalues = {}

        if request.POST:

            for k, v in request.POST.items():
                if k.startswith("question_"):
                    s = k.split("_")
                    if len(s) == 4:
                        #qvalues[s[1]+'_'+v] = '1' # evaluates true in JS
                        if (qvalues.has_key(s[1])):
                            qvalues[s[1]] += " " + v # evaluates true in JS
                        else:
                            qvalues[s[1]] = v # evaluates true in JS
                    elif len(s) == 3 and s[2] == 'comment':
                        qvalues[s[1] + '_' + s[2]] = v
                    else:
                        if (qvalues.has_key(s[1])):
                            qvalues[s[1]] += " " + v
                        else:
                            qvalues[s[1]] = v
                            #print qvalues
            query = convert_qvalues_to_query(qvalues, q_id)
            #print "Query: " + query
            return results_fulltext_aux(request, query)

        qlist_general = []

        for k in qs_list:
            qlist = []
            qs_aux = None
            for question in questions_list[k.id]:
                qs_aux = question.questionset
                #print "Question: " + str(question)
                Type = question.get_type()
                _qnum, _qalpha = split_numal(question.number)

                qdict = {
                    'template': 'questionnaire/%s.html' % (Type),
                    'qnum': _qnum,
                    'qalpha': _qalpha,
                    'qtype': Type,
                    'qnum_class': (_qnum % 2 == 0) and " qeven" or " qodd",
                    'qalpha_class': _qalpha and (ord(_qalpha[-1]) % 2 \
                                                     and ' alodd' or ' aleven') or '',
                }

                # add javascript dependency checks
                cd = question.getcheckdict()
                depon = cd.get('requiredif', None) or cd.get('dependent', None)
                if depon:
                    # extra args to BooleanParser are not required for toString
                    parser = BooleanParser(dep_check)

                    # qdict['checkstring'] = ' checks="%s"' % parser.toString(depon)

                    #It allows only 1 dependency
                    #The line above allows multiple dependencies but it has a bug when is parsing white spaces
                    qdict['checkstring'] = ' checks="dep_check(\'question_%s\')"' % depon

                    qdict['depon_class'] = ' depon_class'
                    jstriggers.append('qc_%s' % question.number)
                    if question.text[:2] == 'h1':
                        jstriggers.append('acc_qc_%s' % question.number)
                if 'default' in cd and not question.number in cookiedict:
                    qvalues[question.number] = cd['default']
                if Type in QuestionProcessors:
                    qdict.update(QuestionProcessors[Type](request, question))
                    if 'jsinclude' in qdict:
                        if qdict['jsinclude'] not in jsinclude:
                            jsinclude.extend(qdict['jsinclude'])
                    if 'cssinclude' in qdict:
                        if qdict['cssinclude'] not in cssinclude:
                            cssinclude.extend(qdict['jsinclude'])
                    if 'jstriggers' in qdict:
                        jstriggers.extend(qdict['jstriggers'])
                        #if 'qvalue' in qdict and not question.number in cookiedict:
                        #    qvalues[question.number] = qdict['qvalue']
                        #

                qlist.append((question, qdict))
            if qs_aux == None:
                #print "$$$$$$ NONE"
                qs_aux = k
            qlist_general.append((qs_aux, qlist))

        errors = {}
        fingerprint_id = generate_hash()
        r = r2r(template_name, request,
                questionset=question_set,
                questionsets=question_set.questionnaire.questionsets,
                runinfo=None,
                errors=errors,
                qlist=qlist,
                progress=None,
                triggers=jstriggers,
                qvalues=qvalues,
                jsinclude=jsinclude,
                cssinclude=cssinclude,
                async_progress=None,
                async_url=None,
                qs_list=qs_list,
                questions_list=qlist_general,
                fingerprint_id=fingerprint_id,
                breadcrumb=True,
        )
        r['Cache-Control'] = 'no-cache'
        r['Expires'] = "Thu, 24 Jan 1980 00:00:00 GMT"

    except:

        raise
    return r



class RequestMonkeyPatch(object):
    POST = {}

    method = POST

    def __init__(self):
        self.POST = {}

    def get_post(self):
        return self.POST


def extract_answers(request2, questionnaire_id, question_set, qs_list):

    question_set2 = question_set
    request = request2
    # Extract files if they exits 
    try:
        if request.FILES:
            for name, f in request.FILES.items():
                handle_uploaded_file(f)
    except:
        pass
        
    qsobjs = QuestionSet.objects.filter(questionnaire=questionnaire_id)
    questionnaire = qsobjs[0].questionnaire
    
    sortid = 0
    
    if request.POST:
        try:
            question_set = request.POST['active_qs']
            sortid = request.POST['active_qs_sortid']
            fingerprint_id = request.POST['fingerprint_id']
        except:
            for qs in qsobjs:
                if qs.sortid == int(sortid):
                    question_set = qs.pk
                    break

    expected = []
    for qset in qsobjs:
        questions = qset.questions()
        for q in questions:
            expected.append(q)

    items = request.POST.items()
    extra = {} # question_object => { "ANSWER" : "123", ... }
    extra_comments = {}
    extra_fields = {}
    # this will ensure that each question will be processed, even if we did not receive
    # any fields for it. Also works to ensure the user doesn't add extra fields in
    for x in expected:
        items.append((u'question_%s_Trigger953' % x.number, None))

    # generate the answer_dict for each question, and place in extra
    for item in items:
        key, value = item[0], item[1]
        if key.startswith('comment_question_'):
            continue
            
        if key.startswith('question_'):
            answer = key.split("_", 2)
            question = get_question(answer[1], questionnaire)
            if not question:
                logging.warn("Unknown question when processing: %s" % answer[1])
                continue
            extra[question] = ans = extra.get(question, {})
            if (len(answer) == 2):
                ans['ANSWER'] = value
            elif (len(answer) == 3):
                ans[answer[2]] = value
            else:
                print "Poorly formed form element name: %r" % answer
                logging.warn("Poorly formed form element name: %r" % answer)
                continue
            extra[question] = ans

            

            comment_id = "comment_question_"+question.number#.replace(".", "")
            try:
                if request.POST and request.POST[comment_id]!='':
                    comment_id_index = "comment_question_"+question.slug
                    extra_comments[question] = request.POST[comment_id]
                    extra_fields[comment_id_index+'_t'] = request.POST[comment_id]
            except KeyError:
                pass
    errors = {}

    print "Extra comments"
    print extra_comments
    
    # Verification of qprocessor answers 
    def verify_answer(question, answer_dict):

        type = question.get_type()

        if "ANSWER" not in answer_dict:
            answer_dict['ANSWER'] = None
        answer = None
        if type in Processors:
            answer = Processors[type](question, answer_dict) or ''
        else:
            raise AnswerException("No Processor defined for question type %s" % type)

        return True

    active_qs_with_errors = False
    
    for question, ans in extra.items():
        
        if u"Trigger953" not in ans:
            logging.warn("User attempted to insert extra question (or it's a bug)")
            continue
        try:
            cd = question.getcheckdict()
            
            depon = cd.get('requiredif', None) or cd.get('dependent', None)
            
            verify_answer(question, ans)
            
        except AnswerException, e:
            errors[question.number] = e

            if (str(question.questionset.id) == question_set):
                #print "active enable"
                active_qs_with_errors = True
        except Exception:
            logging.exception("Unexpected Exception")
            raise

    try:
        questions = question_set2.questions()

        questions_list = {}
        for qset_aux in qs_list:
            questions_list[qset_aux.id] = qset_aux.questions()

        qlist = []
        jsinclude = []      # js files to include
        cssinclude = []     # css files to include
        jstriggers = []
        qvalues = {}

        qlist_general = []

        for k in qs_list:
            qlist = []
            qs_aux = None
            for question in questions_list[k.id]:
                qs_aux = question.questionset
                Type = question.get_type()
                _qnum, _qalpha = split_numal(question.number)
                
                qdict = {
                    'template': 'questionnaire/%s.html' % (Type),
                    'qnum': _qnum,
                    'qalpha': _qalpha,
                    'qtype': Type,
                    'qnum_class': (_qnum % 2 == 0) and " qeven" or " qodd",
                    'qalpha_class': _qalpha and (ord(_qalpha[-1]) % 2 \
                                                     and ' alodd' or ' aleven') or '',
                }

                # add javascript dependency checks
                cd = question.getcheckdict()
                depon = cd.get('requiredif', None) or cd.get('dependent', None)
                if depon:
                    # extra args to BooleanParser are not required for toString
                    parser = BooleanParser(dep_check)

                    # qdict['checkstring'] = ' checks="%s"' % parser.toString(depon)

                    #It allows only 1 dependency
                    #The line above allows multiple dependencies but it has a bug when is parsing white spaces
                    qdict['checkstring'] = ' checks="dep_check(\'question_%s\')"' % depon

                    qdict['depon_class'] = ' depon_class'
                    jstriggers.append('qc_%s' % question.number)
                    if question.text[:2] == 'h1':
                        jstriggers.append('acc_qc_%s' % question.number)
                if 'default' in cd and not question.number in cookiedict:
                    qvalues[question.number] = cd['default']
                if Type in QuestionProcessors:

                    qdict.update(QuestionProcessors[Type](request, question))
                    try:
                        qdict['comment'] = extra_comments[question]
                    except KeyError:
                        pass

                    if 'jsinclude' in qdict:
                        if qdict['jsinclude'] not in jsinclude:
                            jsinclude.extend(qdict['jsinclude'])
                    if 'cssinclude' in qdict:
                        if qdict['cssinclude'] not in cssinclude:
                            cssinclude.extend(qdict['jsinclude'])
                    if 'jstriggers' in qdict:
                        jstriggers.extend(qdict['jstriggers'])
                        
                qlist.append((question, qdict))
                
            if qs_aux == None:
                qs_aux = k
            qlist_general.append((qs_aux, qlist))
    except:
        raise
    return (qlist_general, qlist, jstriggers, qvalues, jsinclude, cssinclude, extra_fields)
    

def database_edit(request, fingerprint_id, questionnaire_id, template_name="database_edit.html"):
    c = CoreEngine()

    results = c.search_fingerprint("id:" + fingerprint_id)
    items = None
    for r in results:
        items = r
        break
    fingerprint_id = r['id']

    users_db = r['user_t']

    created_date = r['created_t']

    try:
        fingerprint_name = r['database_name_t']
    except:
        fingerprint_name = 'unnamed'

    extra = {} 
    # Render page first time.

    request2 = RequestMonkeyPatch()

    request2.method = request.method
    for item in items:
        key = item
        value = items[key]
        if item.startswith("comment_question_"):
            
            slug = item.split("comment_question_")[1]
            results = Slugs.objects.filter(slug1=slug[:-2], question__questionset__questionnaire=questionnaire_id)
            if results == None or len(results) == 0:
                continue
            question = results[0].question
            request2.get_post()['comment_question_%s' % question.number] = value
            continue

        if item == '_version_':
            continue

        results = Slugs.objects.filter(slug1=str(item)[:-2],question__questionset__questionnaire=questionnaire_id )
        print len(results)
        if results == None or len(results) == 0:
            continue
        question = results[0].question
        answer = str(question.number)

        extra[question] = ans = extra.get(question, {})
        if "[" in value:
            value = value.lower().replace("]", "").replace("[", "")
        request2.get_post()['question_%s' % question.number] = value
        
        
        ans['ANSWER'] = value
        
        extra[question] = ans

    errors = {}

    q_id = questionnaire_id
    qs_id = 1

    qs_list = QuestionSet.objects.filter(questionnaire=questionnaire_id)

    question_set = qs_list[int(qs_id)]
    if request.POST:
        (qlist_general, qlist, jstriggers, qvalues, jsinclude, cssinclude, extra_fields) = extract_answers(request, questionnaire_id, question_set, qs_list)
    else:
        (qlist_general, qlist, jstriggers, qvalues, jsinclude, cssinclude, extra_fields) = extract_answers(request2, questionnaire_id, question_set, qs_list)

    if (question_set.sortid == 99 or request.POST):
        # Index on Solr
        try:
            add_city(qlist_general)

            index_answeres_from_qvalues(qlist_general, question_set.questionnaire, users_db,
                                        fingerprint_id, extra_fields=extra_fields, created_date=created_date)
        except:
            raise

    r = r2r(template_name, request,
            questionset=question_set,
            questionsets=question_set.questionnaire.questionsets,
            runinfo=None,
            errors=errors,
            qlist=qlist,
            progress=None,
            triggers=jstriggers,
            qvalues=qvalues,
            jsinclude=jsinclude,
            cssinclude=cssinclude,
            fingerprint_id=fingerprint_id,
            async_progress=None,
            async_url=None,
            qs_list=qs_list,
            questions_list=qlist_general,
            breadcrumb=True,
            name=fingerprint_name.decode('ascii', 'ignore'),
            id=fingerprint_id,
            users_db=users_db,
            created_date=created_date,
            hide_add=True
    )
    r['Cache-Control'] = 'no-cache'
    r['Expires'] = "Thu, 24 Jan 1980 00:00:00 GMT"
    
    return r

class Database:
    id = ''
    name = ''
    date = ''
    last_activity = ''
    


def get_databases_from_db(request):
    user = request.user
    su = Subject.objects.filter(user=user)
    databases = RunInfoHistory.objects.filter(subject=su)

    list_databases = []
    for database in databases:
        database_aux = Database()
        database_aux.id = database.runid
        database_aux.date = database.completed
        answers = Answer.objects.filter(runid=database.runid)
        text = clean_value(str(answers[1].answer))
        info = text[:75] + (text[75:] and '..')
        database_aux.name = info
        list_databases.append(database_aux)
    return list_databases


def get_databases_from_solr(request, query="*:*"):
    c = CoreEngine()
    results = c.search_fingerprint(query)
    print "Solr"
    print results
    list_databases = []
    questionnaires_ids = {}
    qqs = Questionnaire.objects.all()
    for q in qqs:
        questionnaires_ids[q.slug] = (q.pk, q.name)

    for r in results:
        try:
            database_aux = Database()
            
            database_aux.id = r['id']

            if (not r.has_key('created_t')):
                database_aux.date = ''
            else:

                try:
                    database_aux.date = convert_date(r['created_t'])
                except:
                    database_aux.date = ''


            if (not r.has_key('date_last_modification_t')):
                database_aux.date_modification = convert_date(r['created_t'])
            else:

                try:
                    database_aux.date_modification = convert_date(r['date_last_modification_t'])
                except:
                    database_aux.date_modification = convert_date(r['created_t'])

            if (not r.has_key('database_name_t')):
                database_aux.name = '(Unnamed)'
            else:
                database_aux.name = r['database_name_t']

            if (not r.has_key('location_t')):
                database_aux.location = ''
            else:
                database_aux.location = r['location_t']

            if (not r.has_key('institution_name_t')):
                database_aux.institution = ''
            else:
                database_aux.institution = r['institution_name_t']

            if (not r.has_key('contact_administrative_t')):
                database_aux.email_contact = ''
            else:
                database_aux.email_contact = r['contact_administrative_t']

            if (not r.has_key('number_active_patients_jan2012_t')):
                database_aux.number_patients = ''
            else:
                database_aux.number_patients = r['number_active_patients_jan2012_t']

            if (not r.has_key('date_last_modification_t')):
                database_aux.last_activity = ''
            else:
                database_aux.last_activity = r['date_last_modification_t']                
                
            if (not r.has_key('upload-image_t')):
                database_aux.logo = 'nopic.gif'
            else:
                database_aux.logo = r['upload-image_t']

            (ttype, type_name) = questionnaires_ids[r['type_t']]
            database_aux.ttype = ttype
            database_aux.type_name = type_name
            list_databases.append(database_aux)
        except:
            pass
            #raise
    return list_databases


def delete_fingerprint(request, id):
    user = request.user

    c = CoreEngine()
    results = c.search_fingerprint('user_t:' + user.username)
    
    for result in results:
        if (id == result['id']):
            c.delete(id)
            break

    return databases(request)

def force_delete_fingerprint(request, id):
    if not request.user.is_superuser:
        return HttpResponse('Permission denied. Contact administrator of EMIF Catalogue Team', status=403)
    user = request.user
    c = CoreEngine()
    results = c.search_fingerprint('id:' + id)
    
    for result in results:
        if (id == result['id']):
            c.delete(id)
            break

    return databases(request)


def databases(request, page=1, template_name='databases.html'):
    # Get the list of databases for a specific user

    user = request.user
    #list_databases = get_databases_from_db(request)
    _filter = "user_t:" + user.username
    if user.is_superuser:
        _filter = "user_t:*" 
    list_databases = get_databases_from_solr(request, _filter)


    ## Paginator ##
    rows = 5
    myPaginator = Paginator(list_databases, rows)
    try:
        pager =  myPaginator.page(page)
    except PageNotAnInteger, e:
        pager =  myPaginator.page(1)
    ## End Paginator ##

    return render(request, template_name, {'request': request, 'export_my_answers': True,
                                           'list_databases': list_databases, 'breadcrumb': True, 'collapseall': False,
                                           'page_obj': pager,
                                           'api_token': True, 
                                           'owner_fingerprint': False,
                                           'add_databases': True})


def all_databases(request, page=1, template_name='alldatabases.html'):
    #list_databases = get_databases_from_db(request)
    list_databases = get_databases_from_solr(request, "*:*")

    ## Paginator ##
    rows = 5
    myPaginator = Paginator(list_databases, rows)
    try:
        pager =  myPaginator.page(page)
    except PageNotAnInteger, e:
        pager =  myPaginator.page(1)
    ## End Paginator ##
    
    return render(request, template_name, {'request': request, 'export_all_answers': True, 'data_table': True,
                                           'list_databases': list_databases,
                                            'breadcrumb': True, 'collapseall': False, 
                                            'geo': True,
                                            'page_obj': pager,
                                            'add_databases': True})

def qs_data_table(request, template_name='qs_data_table.html'):
    db_type = request.POST.get("db_type")
    qset = request.POST.get("qset")
    
    answers = []
    # get only databases with correct type
    list_databases = get_databases_from_solr(request, "type_t:"+re.sub(r'\s+', '', db_type.lower()))
    titles = []
    
    for t in list_databases:

        if t.type_name == db_type:
            qsets, name = createqsets(t.id)
            
            q_list = []
            a_list = []            
            for group in qsets.ordered_items():

                (k, qs) = group    
              
                if k == qset:
                    for q in qs.list_ordered_tags:
                        q_list.append(q)
                        a_list.append(q.value)
                continue
            titles = ('Name', (q_list))
            

            answers.append((name, (a_list)))
                # print answers
            
            
    return render(request, template_name, {'request': request, 'export_all_answers': True, 'breadcrumb': False, 'collapseall': False, 'geo': False, 'titles': titles, 'answers': answers})

def all_databases_data_table(request, template_name='alldatabases_data_table.html'):
    answers = []
    list_databases = get_databases_from_solr(request, "*:*")
    titles = []
    
    #dictionary of database types
    databases_types = {}

    if list_databases:
        # Creating list of database types
        for t in list_databases:
            if not t.type_name in databases_types:
                qsets, name = createqsets(t.id)
                
                databases_types[t.type_name] = qsets.ordered_items()  
            
            '''    this code was loading all the qsets of all the dbs etc etctera
            id = t.id
            qsets, name, db_owners, fingerprint_ttype = createqsets(id)
            q_list = []
            for group in qsets.ordered_items():
                (k, qs) = group
                for q in qs.list_ordered_tags:
                    q_list.append(q)
            titles = ('Name', (q_list))
            a_list = []
            for group in qsets.ordered_items():
                (k, qs) = group
                for q in qs.list_ordered_tags:
                    a_list.append(q.value)
            answers.append((name, (a_list)))
                # print answers

        # since we dont have access to the structure directly (?)
        # i use a random id for each type to get the qset types ?
        if databases_types:
            #print databases_types['adcohort'][0].id
            for type in databases_types:
                qsets, name = createqsets(databases_types[type][0].id)
                
                qsets_by_type[type] = qsets.ordered_items()
 '''           
    # print titles
    # print answers

    return render(request, template_name, {'request': request, 'export_all_answers': True, 'titles': titles,
                                           'answers': answers, 'breadcrumb': True, 'collapseall': False, 'geo': True,
                                           'list_databases': list_databases,
                                           'databases_types': databases_types
                                           })


def createqsets(runcode, qsets=None, clean=True):
    #print "createqsets"
    c = CoreEngine()
    results = c.search_fingerprint('id:' + runcode)

    if qsets == None:
        qsets = ordered_dict()
    name = ""
    list_values = []
    blacklist = ['created_t', 'type_t', '_version_', 'date_last_modification_t']
    name = "Not defined."
    users = ""
    fingerprint_ttype = ""

    db_owners = "" 


    questionnaires_ids = {}
    qqs = Questionnaire.objects.all()
    for q in qqs:
        questionnaires_ids[q.slug] = (q.pk, q.name)


    for result in results:


        (fingerprint_ttype, type_name) = questionnaires_ids[result['type_t']]

        # Get the slug of fingerprint type
        q_aux = Questionnaire.objects.filter(slug=result['type_t'])

        try:
            users = result['user_t']
            db_owners = result['user_t']
        except:
            pass

        list_qsets = QuestionSet.objects.filter(questionnaire=q_aux[0]).order_by('sortid')

        
        for qset in list_qsets:

            if (qset.sortid != 0 and qset.sortid != 99):
                question_group = QuestionGroup()
                question_group.sortid = qset.sortid
                
                qsets[qset.text] = question_group
                qset.sortid
                list_questions = Question.objects.filter(questionset=qset).order_by('number')
                for question in list_questions:
                    t = Tag()
                    t.tag = question.text.encode('utf-8')
                    t.value = ""
                    t.number = question.number
                    t.ttype = question.type
                    question_group.list_ordered_tags.append(t)



                qsets[qset.text] = question_group

        for k in result:
            #print k
            if k in blacklist:
                continue
            if k.startswith("comment_question_"):
                continue

            t = Tag()

            aux_results = Slugs.objects.filter(slug1=k[:-2], question__questionset__questionnaire=q_aux[0].pk)
            qs = None
            question_group = None
            q_number = None
            if len(aux_results) > 0:
                text = aux_results[0].description
                qs = aux_results[0].question.questionset.text
                q_number = qs = aux_results[0].question.number
                if qsets.has_key(aux_results[0].question.questionset.text):
                    # Add the Tag to the QuestionGroup
                    question_group = qsets[aux_results[0].question.questionset.text]
                else:
                    # Add a new QuestionGroup
                    question_group = QuestionGroup()
                    qsets[aux_results[0].question.questionset.text] = question_group
                    
            else:
                text = k

            info = text
            t.tag = info
            #print t.tag

            if question_group != None and question_group.list_ordered_tags != None:
                try:
                    t = question_group.list_ordered_tags[question_group.list_ordered_tags.index(t)]
                except:
                    pass

            value = clean_value(str(result[k].encode('utf-8')))
            
            try:

               t.comment = result['comment_question_'+k]
               #print t.comment
            except KeyError:
               pass
            if clean:
                t.value = value.replace("#", " ")
            else:
                t.value = value

            if k == "database_name_t":
                name = t.value
            list_values.append(t)
            if question_group != None:
                try:
                    question_group.list_ordered_tags[question_group.list_ordered_tags.index(t)] = t
                except:
                    pass
        break
    # What should I do with this code?
    # I know that it actually do nothing    
    if (users!=""):
        users.split(" \\ ")

    return (qsets, name, db_owners, fingerprint_ttype)

   
# TODO: move to another place, maybe API? 
def get_api_info(fingerprint_id):
    """This is an auxiliar method to get the API Info
    """
    result = {}

    
    results = FingerprintAPI.objects.filter(fingerprintID=fingerprint_id)
    result = {}
    for r in results:
        result[r.field] = r.value
    return result


def fingerprint(request, runcode, qs, template_name='database_info.html'):
    
    
    qsets, name, db_owners, fingerprint_ttype = createqsets(runcode)

    if fingerprint_ttype == "":
        raise "There is missing ttype of questionarie, something is really wrong"

    apiinfo = json.dumps(get_api_info(runcode));

    owner_fingerprint = False
    for owner in db_owners.split(" "):
        print owner
        print request.user.username
        if (owner == request.user.username):
            owner_fingerprint = True
    
    return render(request, template_name, 
        {'request': request, 'qsets': qsets, 'export_bd_answers': True, 
        'apiinfo': apiinfo, 'fingerprint_id': runcode,
                   'breadcrumb': True, 'breadcrumb_name': name.decode('ascii', 'ignore'),
                    'style': qs, 'collapseall': False, 
                    'owner_fingerprint':owner_fingerprint,
                    'fingerprint_dump': True,
                    'fingerprint_ttype': fingerprint_ttype,
                    })


def get_questionsets_list(runinfo):
    # Get questionnaire
    current = runinfo.questionset
    sets = current.questionnaire.questionsets()
    return sets


def show_full_questionnaire_ro(request, qs_id, runinfo, errors={},
                               reverse_name='questionaries_with_sets'):
    """
    Return the QuestionSet template

    Also add the javascript dependency code.
    """

    #r = assure_authenticated_or_redirect(request)

    #if r:
    #    return r
    questionnaire_id = runinfo
    qu = get_object_or_404(Questionnaire, id=questionnaire_id)

    #qs = qu.questionsets()
    qs = qu.questionsets()[0]
    user = request.user
    su = Subject.objects.filter(user=user)

    if su:
        su = su[0]
    else:
        su = Subject(user=user,
                     first_name=user.first_name,
                     last_name=user.last_name,
                     email=user.email,
                     state='active')
        su.save()

    qs = qs_id
    try:
        qsobj = QuestionSet.objects.filter(pk=qs)[0]
    except:
        pass

    questionnaire = qsobj.questionnaire
    questionset = qsobj

    expected = questionset.questions()

    items = request.POST.items()
    extra = {} # question_object => { "ANSWER" : "123", ... }

    # this will ensure that each question will be processed, even if we did not receive
    # any fields for it. Also works to ensure the user doesn't add extra fields in
    for x in expected:
        items.append((u'question_%s_Trigger953' % x.number, None))

    # generate the answer_dict for each question, and place in extra
    for item in items:
        key, value = item[0], item[1]
        print key
        print value
        if key.startswith('question_'):
            answer = key.split("_", 2)
            question = get_question(answer[1], questionnaire)
            if not question:
                logging.warn("Unknown question when processing: %s" % answer[1])
                continue
            extra[question] = ans = extra.get(question, {})
            print answer
            if (len(answer) == 2):
                ans['ANSWER'] = value
            elif (len(answer) == 3):
                ans[answer[2]] = value
            else:
                logging.warn("Poorly formed form element name: %r" % answer)
                continue
            extra[question] = ans

    errors = {}
    for question, ans in extra.items():
        if not question_satisfies_checks(question, runinfo):
            continue
        if u"Trigger953" not in ans:
            logging.warn("User attempted to insert extra question (or it's a bug)")
            continue
        try:
            cd = question.getcheckdict()
            # requiredif is the new way
            depon = cd.get('requiredif', None) or cd.get('dependent', None)
            if depon:
                depparser = BooleanParser(dep_check, runinfo, extra)
                if not depparser.parse(depon):
                    # if check is not the same as answer, then we don't care
                    # about this question plus we should delete it from the DB
                    delete_answer(question, runinfo.subject, runinfo.runid)
                    if cd.get('store', False):
                        runinfo.set_cookie(question.number, None)
                    continue
            add_answer(runinfo, question, ans)
            if cd.get('store', False):
                runinfo.set_cookie(question.number, ans['ANSWER'])
        except AnswerException, e:
            errors[question.number] = e
        except Exception:
            logging.exception("Unexpected Exception")
            transaction.rollback()
            raise

    if len(errors) > 0:
        res = show_questionnaire(request, runinfo, errors=errors)
        return res

    next = questionset.next()

    if next is None: # we are finished
        return finish_questionnaire(runinfo, questionnaire)

    return redirect_to_qs(runinfo)


def show_full_questionnaire(request, runinfo, errors={},
                            reverse_name='questionaries_with_sets'):
    """
    Return the QuestionSet template

    Also add the javascript dependency code.
    """

    
    questionnaire_id = runinfo
    qu = get_object_or_404(Questionnaire, id=questionnaire_id)

    #qs = qu.questionsets()
    qs = qu.questionsets()[0]
    user = request.user
    su = Subject.objects.filter(user=user)

    if su:
        su = su[0]
    else:
        su = Subject(user=user,
                     first_name=user.first_name,
                     last_name=user.last_name,
                     email=user.email,
                     state='active')
        su.save()

    hash = md5.new()
    hash.update("".join(map(lambda i: chr(random.randint(0, 255)), range(16))))
    hash.update(settings.SECRET_KEY)
    key = hash.hexdigest()
    run = RunInfo(subject=su, random=key, runid=key, questionset=qs)
    run.save()
    questionsets = run.questionset.questionnaire.questionsets()

    return HttpResponseRedirect(reverse(reverse_name,
                                        kwargs={'runcode': key}))


def redirect_to_qs(runinfo):
    "Redirect to the correct and current questionset URL for this RunInfo"

    # cache current questionset
    qs = runinfo.questionset

    # skip questionsets that don't pass
    if not questionset_satisfies_checks(runinfo.questionset, runinfo):

        next = runinfo.questionset.next()

        while next and not questionset_satisfies_checks(next, runinfo):
            next = next.next()

        runinfo.questionset = next
        runinfo.save()

        hasquestionset = bool(next)
    else:
        hasquestionset = True

    # empty ?
    if not hasquestionset:
        logging.warn('no questionset in questionnaire which passes the check')
        return finish_questionnaire(runinfo, qs.questionnaire)

    url = reverse("questionset_sets",
                  args=[runinfo.random, runinfo.questionset.sortid])

    print "redirect to url: " + url
    return HttpResponseRedirect(url)


@transaction.commit_on_success
def questionaries_with_sets(request, runcode=None, qs=None, template_name='database_edit.html'):
    print "questionaries_with_sets"
    print "RunCode: " + str(runcode)
    print "Qs: " + str(qs)
    # if runcode provided as query string, redirect to the proper page
    if not runcode:
        runcode = request.GET.get('runcode')
        print "RunCode inside: " + runcode
        if not runcode:
            return HttpResponseRedirect("/")
        else:

            __qs = int(qs) + 1
            __qs += 1
            print HttpResponseRedirect(reverse("questionaries_with_sets", args=[runcode, str(__qs)]))
            return HttpResponseRedirect(reverse("questionaries_with_sets", args=[runcode, str(__qs)]))

    runinfo = get_runinfo(runcode)
    print "Runinfo:" + str(runinfo)

    if not runinfo:
        transaction.commit()
        return HttpResponseRedirect('/')


    #qs = runinfo.questionset
    if request.method != "POST":
        if qs is not None:
            qs = get_object_or_404(QuestionSet, sortid=qs, questionnaire=runinfo.questionset.questionnaire)
            print "qs.sortid=" + str(qs.sortid)
            print "runinfor.qs.sortid" + str(runinfo.questionset.sortid)
            if runinfo.random.startswith('test:'):
                pass # ok for testing
            elif qs.sortid > runinfo.questionset.sortid:
                # you may jump back, but not forwards
                return redirect_to_qs(runinfo)
            runinfo.questionset = qs
            runinfo.save()
            transaction.commit()
            # no questionset id in URL, so redirect to the correct URL
        if qs is None:
            return redirect_to_qs(runinfo)
        return show_fingerprint_page(request, runinfo)

    # Get several questionsets
    print "RunInfo: " + str(runinfo)
    print "runcode: " + str(runcode)
    print "QS: " + str(qs)

    # -------------------------------------
    # --- Process POST with QuestionSet ---
    # -------------------------------------

    # if the submitted page is different to what runinfo says, update runinfo
    # XXX - do we really want this?
    qs = request.POST.get('questionset_id', None)
    try:
        qsobj = QuestionSet.objects.filter(pk=qs)[0]
        if qsobj.questionnaire == runinfo.questionset.questionnaire:
            if runinfo.questionset != qsobj:
                runinfo.questionset = qsobj
                runinfo.save()
    except:
        pass

    questionnaire = runinfo.questionset.questionnaire
    questionset = runinfo.questionset

    # to confirm that we have the correct answers
    expected = questionset.questions()

    items = request.POST.items()
    extra = {} # question_object => { "ANSWER" : "123", ... }

    # this will ensure that each question will be processed, even if we did not receive
    # any fields for it. Also works to ensure the user doesn't add extra fields in
    for x in expected:
        items.append((u'question_%s_Trigger953' % x.number, None))

    # generate the answer_dict for each question, and place in extra
    for item in items:
        key, value = item[0], item[1]
        if key.startswith('question_'):
            answer = key.split("_", 2)
            question = get_question(answer[1], questionnaire)
            if not question:
                logging.warn("Unknown question when processing: %s" % answer[1])
                continue
            extra[question] = ans = extra.get(question, {})
            if (len(answer) == 2):
                ans['ANSWER'] = value
            elif (len(answer) == 3):
                ans[answer[2]] = value
            else:
                logging.warn("Poorly formed form element name: %r" % answer)
                continue
            extra[question] = ans

    errors = {}
    for question, ans in extra.items():
        if not question_satisfies_checks(question, runinfo):
            continue
        if u"Trigger953" not in ans:
            logging.warn("User attempted to insert extra question (or it's a bug)")
            continue
        try:
            cd = question.getcheckdict()
            # requiredif is the new way
            depon = cd.get('requiredif', None) or cd.get('dependent', None)
            if depon:
                depparser = BooleanParser(dep_check, runinfo, extra)
                if not depparser.parse(depon):
                    # if check is not the same as answer, then we don't care
                    # about this question plus we should delete it from the DB
                    delete_answer(question, runinfo.subject, runinfo.runid)
                    if cd.get('store', False):
                        runinfo.set_cookie(question.number, None)
                    continue
            add_answer(runinfo, question, ans)
            if cd.get('store', False):
                runinfo.set_cookie(question.number, ans['ANSWER'])
        except AnswerException, e:
            errors[question.number] = e
        except Exception:
            logging.exception("Unexpected Exception")
            transaction.rollback()
            raise

    if len(errors) > 0:
        res = show_questionnaire(request, runinfo, errors=errors)
        transaction.rollback()
        return res

    questionset_done.send(sender=None, runinfo=runinfo, questionset=questionset)

    next = questionset.next()
    while next and not questionset_satisfies_checks(next, runinfo):
        next = next.next()
    runinfo.questionset = next
    runinfo.save()

    if next is None: # we are finished
        return finish_questionnaire(runinfo, questionnaire)

    transaction.commit()
    return redirect_to_qs(runinfo)


def next_questionset_order_by_sortid(questionset_id, questionnaire_id):
    qsobjs = QuestionSet.objects.filter(questionnaire=questionnaire_id).order_by('sortid')
    next = False
    next_qs = None
    for qs in qsobjs:
        #print qs.pk
        #print questionset_id
        if next:
            next_qs = qs
            next = False
            break
        if qs.pk == int(questionset_id):
            next = True
    return next_qs

def handle_uploaded_file(f):
    print "abspath"

    with open(os.path.join(os.path.abspath(settings.PROJECT_DIR_ROOT + 'emif/static/upload_images/'), f.name),
              'wb+') as destination:
        for chunk in f.chunks():
            destination.write(chunk)


def check_database_add_conditions(request, questionnaire_id, sortid,
                                  template_name='database_add.html', users_db=None, created_date=None):
    # -------------------------------------
    # --- Process POST with QuestionSet ---
    # -------------------------------------

    print "check_database_add_conditions"
    try:
        if request.FILES:
            print "file upload:"
            for name, f in request.FILES.items():
                handle_uploaded_file(f)
    except:
        raise

    qsobjs = QuestionSet.objects.filter(questionnaire=questionnaire_id)
    questionnaire = qsobjs[0].questionnaire
    question_set = None

    try:
        question_set = request.POST['active_qs']
        sortid = request.POST['active_qs_sortid']
    except:
        for qs in qsobjs:
            if qs.sortid == int(sortid):
                question_set = qs.pk
                break

    fingerprint_id = request.POST['fingerprint_id']
    
    return show_fingerprint_page_errors(request, questionnaire_id, question_set,
                                        errors={}, template_name='database_add.html', next=True, sortid=sortid,
                                        fingerprint_id=fingerprint_id, users_db=users_db, created_date=created_date)


def show_fingerprint_page_errors(request, q_id, qs_id, errors={}, template_name='database_add.html',
                                 next=False, sortid=0, fingerprint_id=None, users_db=None, created_date=None):
    """
    Return the QuestionSet template

    Also add the javascript dependency code.
    """
    try:

        qs_list = QuestionSet.objects.filter(questionnaire=q_id).order_by('sortid')
        
        initial_sort = sortid

        if (int(sortid) == 99):
            sortid = len(qs_list) - 1
        question_set = qs_list[int(sortid)]
        
        questions = question_set.questions()
        
        questions_list = {}
        for qset_aux in qs_list:
            questions_list[qset_aux.id] = qset_aux.questions()
        
        qlist = []
        jsinclude = []      # js files to include
        cssinclude = []     # css files to include
        jstriggers = []
        qvalues = {}

        qlist_general = []
        extra_fields = {}
        for k in qs_list:
            qlist = []
            qs_aux = None
            
            for question in questions_list[k.id]:
                
                qs_aux = question.questionset
                
                Type = question.get_type()
                _qnum, _qalpha = split_numal(question.number)

                qdict = {
                    'template': 'questionnaire/%s.html' % (Type),
                    'qnum': _qnum,
                    'qalpha': _qalpha,
                    'qtype': Type,
                    'qnum_class': (_qnum % 2 == 0) and " qeven" or " qodd",
                    'qalpha_class': _qalpha and (ord(_qalpha[-1]) % 2 \
                                                     and ' alodd' or ' aleven') or '',
                }

                # add javascript dependency checks
                cd = question.getcheckdict()
                depon = cd.get('requiredif', None) or cd.get('dependent', None)
                if depon:
                    # extra args to BooleanParser are not required for toString
                    parser = BooleanParser(dep_check)
                    # qdict['checkstring'] = ' checks="%s"' % parser.toString(depon)

                    #It allows only 1 dependency
                    #The line above allows multiple dependencies but it has a bug when is parsing white spaces
                    qdict['checkstring'] = ' checks="dep_check(\'question_%s\')"' % depon

                    qdict['depon_class'] = ' depon_class'
                    jstriggers.append('qc_%s' % question.number)
                    if question.text[:2] == 'h1':
                        jstriggers.append('acc_qc_%s' % question.number)
                if 'default' in cd and not question.number in cookiedict:
                    qvalues[question.number] = cd['default']
                if Type in QuestionProcessors:
                    qdict.update(QuestionProcessors[Type](request, question))
                    if 'jsinclude' in qdict:
                        if qdict['jsinclude'] not in jsinclude:
                            jsinclude.extend(qdict['jsinclude'])
                    if 'cssinclude' in qdict:
                        if qdict['cssinclude'] not in cssinclude:
                            cssinclude.extend(qdict['jsinclude'])
                    if 'jstriggers' in qdict:
                        jstriggers.extend(qdict['jstriggers'])
                        #if 'qvalue' in qdict and not question.number in cookiedict:
                        #    qvalues[question.number] = qdict['qvalue']

                qlist.append((question, qdict))
                comment_id = "comment_question_"+question.number#.replace(".", "")
                if request.POST and request.POST[comment_id]!='':
                    comment_id_index = "comment_question_"+question.slug
                    extra_fields[comment_id_index+'_t'] = request.POST[comment_id]
                    qdict['comment'] = request.POST[comment_id]

            
            if qs_aux == None:
                qs_aux = k
            qlist_general.append((qs_aux, qlist))
        
        #print "Extra fields : " + str(extra_fields)
        if (fingerprint_id != None):

            if users_db==None:
                users_db = request.user.username

            # adding city to cities database (if doesnt exist)
            add_city(qlist_general)

            index_answeres_from_qvalues(qlist_general, question_set.questionnaire, users_db,
                                        fingerprint_id, extra_fields=extra_fields, created_date=created_date)

        r = r2r(template_name, request,
                questionset=question_set,
                questionsets=question_set.questionnaire.questionsets,
                runinfo=None,
                errors=errors,
                qlist=qlist,
                progress=None,
                triggers=jstriggers,
                qvalues=qvalues,
                jsinclude=jsinclude,
                cssinclude=cssinclude,
                async_progress=None,
                async_url=None,
                qs_list=qs_list,
                questions_list=qlist_general,
                fingerprint_id=fingerprint_id,
                breadcrumb=True,
                extra_fields=extra_fields
        )
        r['Cache-Control'] = 'no-cache'
        r['Expires'] = "Thu, 24 Jan 1980 00:00:00 GMT"
    except:
        raise
    return r





def show_fingerprint_page_read_only(request, q_id, qs_id, SouMesmoReadOnly=False, errors={}, template_name='advanced_search.html'):

    """
    Return the QuestionSet template

    Also add the javascript dependency code.
    """
    if template_name == "database_add.html" :
        hide_add = True
    else:
        hide_add = False
    
    try:

        qs_list = QuestionSet.objects.filter(questionnaire=q_id).order_by('sortid')

        #print "Q_id: " + q_id
        #print "Qs_id: " + qs_id
        #print "QS List: " + str(qs_list)
        if (int(qs_id) == 99):
            qs_id = len(qs_list) - 1
        question_set = qs_list[int(qs_id)]
        #questions = Question.objects.filter(questionset=qs_id)

        questions = question_set.questions()
        #print "Questions: " + str(questions)
        #print "QuestionSet: " + str(question_set)

        questions_list = {}
        for qset_aux in qs_list:
            #questions_aux = Question.objects.filter(questionset=qset_aux)
            questions_list[qset_aux.id] = qset_aux.questions()
            #print "here"

        qlist = []
        jsinclude = []      # js files to include
        cssinclude = []     # css files to include
        jstriggers = []
        qvalues = {}
        if request.POST:
            for k, v in request.POST.items():
                
                if (len(v)==0):
                    continue
                if k.startswith("question_"):
                    s = k.split("_")
                    if len(s) == 4:
                        #qvalues[s[1]+'_'+v] = '1' # evaluates true in JS
                        if (qvalues.has_key(s[1])):
                            qvalues[s[1]] += " " + v # evaluates true in JS
                        else:
                            qvalues[s[1]] = v # evaluates true in JS
                    elif len(s) == 3 and s[2] == 'comment':
                        qvalues[s[1] + '_' + s[2]] = v
                    else:
                        if (qvalues.has_key(s[1])):
                            qvalues[s[1]] += " " + v
                        else:
                            qvalues[s[1]] = v
                            #print qvalues
            query = convert_qvalues_to_query(qvalues, q_id)
            print "Query: " + query
            return results_fulltext_aux(request, query)

        qlist_general = []

        for k in qs_list:
            qlist = []
            qs_aux = None
            for question in questions_list[k.id]:
                qs_aux = question.questionset
                #print "Question: " + str(question)
                Type = question.get_type()
                if SouMesmoReadOnly and Type == 'open-button':
                   Type = "open"
               
                _qnum, _qalpha = split_numal(question.number)

                qdict = {
                    'template': 'questionnaire/%s.html' % (Type),
                    'qnum': _qnum,
                    'qalpha': _qalpha,
                    'qtype': Type,
                    'qnum_class': (_qnum % 2 == 0) and " qeven" or " qodd",
                    'qalpha_class': _qalpha and (ord(_qalpha[-1]) % 2 \
                                                     and ' alodd' or ' aleven') or '',
                }

                # add javascript dependency checks
                cd = question.getcheckdict()
                depon = cd.get('requiredif', None) or cd.get('dependent', None)
                if depon:
                    # extra args to BooleanParser are not required for toString
                    parser = BooleanParser(dep_check)

                    # qdict['checkstring'] = ' checks="%s"' % parser.toString(depon)

                    #It allows only 1 dependency
                    #The line above allows multiple dependencies but it has a bug when is parsing white spaces
                    qdict['checkstring'] = ' checks="dep_check(\'question_%s\')"' % depon

                    qdict['depon_class'] = ' depon_class'
                    jstriggers.append('qc_%s' % question.number)
                    if question.text[:2] == 'h1':
                        jstriggers.append('acc_qc_%s' % question.number)
                if 'default' in cd and not question.number in cookiedict:
                    qvalues[question.number] = cd['default']
                if Type in QuestionProcessors:
                    qdict.update(QuestionProcessors[Type](request, question))
                    if 'jsinclude' in qdict:
                        if qdict['jsinclude'] not in jsinclude:
                            jsinclude.extend(qdict['jsinclude'])
                    if 'cssinclude' in qdict:
                        if qdict['cssinclude'] not in cssinclude:
                            cssinclude.extend(qdict['jsinclude'])
                    if 'jstriggers' in qdict:
                        jstriggers.extend(qdict['jstriggers'])
                        #if 'qvalue' in qdict and not question.number in cookiedict:
                        #    qvalues[question.number] = qdict['qvalue']
                        #

                qlist.append((question, qdict))
                
            if qs_aux == None:
                #print "$$$$$$ NONE"
                qs_aux = k
            qlist_general.append((qs_aux, qlist))

        errors = {}
        fingerprint_id = generate_hash()
        r = r2r(template_name, request,
                questionset=question_set,
                questionsets=question_set.questionnaire.questionsets,
                runinfo=None,
                errors=errors,
                qlist=qlist,
                progress=None,
                triggers=jstriggers,
                qvalues=qvalues,
                jsinclude=jsinclude,
                cssinclude=cssinclude,
                async_progress=None,
                async_url=None,
                qs_list=qs_list,
                questions_list=qlist_general,
                fingerprint_id=fingerprint_id,
                breadcrumb=True,
                hide_add = hide_add,
        )
        r['Cache-Control'] = 'no-cache'
        r['Expires'] = "Thu, 24 Jan 1980 00:00:00 GMT"

    except:

        raise
    return r


def feedback(request, template_name='feedback.html'):

    if request.method == 'POST':  # If the form has been submitted...
        form = ContactForm(request.POST)
        if form.is_valid():  # All validation rules pass

            subject = request.POST.get('topic', '').encode('ascii', 'ignore')
            name = request.POST.get('name', '').encode('ascii', 'ignore')
            message = request.POST.get('message', '').encode('ascii', 'ignore')
            from_email = request.POST.get('email', '')

            emails_to_feedback = []
            for k, v in settings.ADMINS:
                emails_to_feedback.append(v)

            try:
                message_admin = "Name: " + str(name) + "\nEmail: " + from_email + "\n\nMessage:\n" + str(message)
                message = "Dear " + name + ",\n\nThank you for giving us your feedback.\n\nYour message will be analyzed by EMIF Catalogue team.\n\nMessage sent:\n" + str(message) + "\n\nSincerely,\nEMIF Catalogue"
                # Send email to admins
                send_mail(subject, message_admin, settings.DEFAULT_FROM_EMAIL, emails_to_feedback)
                # Send email to user with the copy of feedback message
                send_mail(subject, message, settings.DEFAULT_FROM_EMAIL, [from_email])

            except BadHeaderError:
                return HttpResponse('Invalid header found.')

            return feedback_thankyou(request)

    else:
        form = ContactForm()  # An unbound form
    return render(request, template_name, {'form': form, 'request': request, 'breadcrumb': True})

        # return render_to_response('feedback.html', {'form': ContactForm()},
        #     RequestContext(request))


def feedback_thankyou(request, template_name='feedback_thankyou.html'):
    return render(request, template_name, {'request': request, 'breadcrumb': True})



def show_fingerprint_page(request, runinfo, errors={}, template_name='database_edit.html'):
    """
    Return the QuestionSet template

    Also add the javascript dependency code.
    """
    questions = runinfo.questionset.questions()

    questions = runinfo.questionset.questions()

    qlist = []
    jsinclude = []      # js files to include
    cssinclude = []     # css files to include
    jstriggers = []
    qvalues = {}

    # initialize qvalues        
    cookiedict = runinfo.get_cookiedict()
    for k, v in cookiedict.items():
        qvalues[k] = v

    substitute_answer(qvalues, runinfo.questionset)

    for question in questions:

        # if we got here the questionset will at least contain one question
        # which passes, so this is all we need to check for
        if not question_satisfies_checks(question, runinfo):
            continue

        Type = question.get_type()
        _qnum, _qalpha = split_numal(question.number)

        qdict = {
            'template': 'questionnaire/%s.html' % (Type),
            'qnum': _qnum,
            'qalpha': _qalpha,
            'qtype': Type,
            'qnum_class': (_qnum % 2 == 0) and " qeven" or " qodd",
            'qalpha_class': _qalpha and (ord(_qalpha[-1]) % 2 \
                                             and ' alodd' or ' aleven') or '',
        }

        # substitute answer texts
        substitute_answer(qvalues, question)

        # add javascript dependency checks
        cd = question.getcheckdict()
        depon = cd.get('requiredif', None) or cd.get('dependent', None)
        if depon:
            # extra args to BooleanParser are not required for toString
            parser = BooleanParser(dep_check)
            # qdict['checkstring'] = ' checks="%s"' % parser.toString(depon)

            #It allows only 1 dependency
            #The line above allows multiple dependencies but it has a bug when is parsing white spaces
            qdict['checkstring'] = ' checks="dep_check(\'question_%s\')"' % depon

            qdict['depon_class'] = ' depon_class'
            jstriggers.append('qc_%s' % question.number)
            if question.text[:2] == 'h1':
                jstriggers.append('acc_qc_%s' % question.number)
        if 'default' in cd and not question.number in cookiedict:
            qvalues[question.number] = cd['default']
        if Type in QuestionProcessors:
            qdict.update(QuestionProcessors[Type](request, question))
            if 'jsinclude' in qdict:
                if qdict['jsinclude'] not in jsinclude:
                    jsinclude.extend(qdict['jsinclude'])
            if 'cssinclude' in qdict:
                if qdict['cssinclude'] not in cssinclude:
                    cssinclude.extend(qdict['jsinclude'])
            if 'jstriggers' in qdict:
                jstriggers.extend(qdict['jstriggers'])
            if 'qvalue' in qdict and not question.number in cookiedict:
                qvalues[question.number] = qdict['qvalue']

        qlist.append((question, qdict))

    try:
        has_progress = settings.QUESTIONNAIRE_PROGRESS in ('async', 'default')
        async_progress = settings.QUESTIONNAIRE_PROGRESS == 'async'
    except AttributeError:
        has_progress = True
        async_progress = False

    if has_progress:
        if async_progress:
            progress = cache.get('progress' + runinfo.random, 1)
        else:
            progress = get_progress(runinfo)
    else:
        progress = 0

    if request.POST:
        for k, v in request.POST.items():
            if k.startswith("question_"):
                s = k.split("_")
                if len(s) == 4:
                    qvalues[s[1] + '_' + v] = '1' # evaluates true in JS
                elif len(s) == 3 and s[2] == 'comment':
                    qvalues[s[1] + '_' + s[2]] = v
                else:
                    qvalues[s[1]] = v
    errors = {}
    r = r2r(template_name, request,
            questionset=runinfo.questionset,
            questionsets=runinfo.questionset.questionnaire.questionsets,
            runinfo=runinfo,
            errors=errors,
            qlist=qlist,
            progress=progress,
            triggers=jstriggers,
            qvalues=qvalues,
            jsinclude=jsinclude,
            cssinclude=cssinclude,
            async_progress=async_progress,
            async_url=reverse('progress', args=[runinfo.random]),
            breadcrumb=True
    )
    r['Cache-Control'] = 'no-cache'
    r['Expires'] = "Thu, 24 Jan 1980 00:00:00 GMT"
    return r


def create_auth_token(request, page=1, templateName='api-key.html'):
    """
    Method to create token to authenticate when calls REST API
    """

    user = request.user
    if not Token.objects.filter(user=user).exists():
        token = Token.objects.create(user=request.user)
    else:
        token = Token.objects.get(user=user)

    # print token

    list_databases = get_databases_from_solr(request, "user_t:" + user.username)
    # for database in list_databases:
    #     print database.id

     ## Paginator ##
    rows = 5
    myPaginator = Paginator(list_databases, rows)
    try:
        pager =  myPaginator.page(page)
    except PageNotAnInteger, e:
        pager =  myPaginator.page(1)
    ## End Paginator ##

    return render_to_response(templateName, {'list_databases': list_databases, 'token': token, 'user': user,
                              'request': request, 'breadcrumb': True, 'page_obj': pager}, RequestContext(request))


def sharedb(request, db_id, template_name="sharedb.html"):
    if not request.method == 'POST':
        return HttpResponse('Invalid header found. The request need to be POST')

    # Verify if it is a valid email
    email = request.POST.get('email', '')
    if (email == None or email==''):
        return HttpResponse('Invalid email address.')

    # Verify if it is a valid user name
    username_to_share = User.objects.get(email__exact=email)

    if (username_to_share==None):
        return HttpResponse('Invalid username.')


    # Verify if it is a valid database 
    if (db_id == None or db_id==''):
        return HttpResponse('Invalid email address.')  
    c = CoreEngine()
    results = c.search_fingerprint('id:' + db_id)
    if (len(results)!=1):
        return HttpResponse('Invalid database identifier.')  

    subject = "EMIF Catalogue: A new database has been shared with you."
    name = username_to_share.get_full_name()
    message = request.POST.get('message', '')
    from_email = request.POST.get('email', '')
    # import pdb
    # pdb.set_trace()
    __objs = SharePending.objects.filter(db_id=db_id, pending=True, user=username_to_share)
    if (len(__objs)>0):
        return HttpResponse('Already contains databases')  

    share_pending = SharePending()
    share_pending.user = username_to_share
    share_pending.db_id = db_id
    share_pending.activation_code = generate_hash()
    share_pending.pending = True
    share_pending.user_invite = request.user 
    share_pending.save()

    link_activation = settings.BASE_URL + "share/activation/"+share_pending.activation_code

    emails_to_feedback = []
    for k, v in settings.ADMINS:
        emails_to_feedback.append(v)

    try:
        
        message = """Dear %s,\n\n
            \n\n
            %s has shared a new database with you. 
            Now you're able to edit and manage the database. \n\n
            To activate the database in your account, please open this link:
            %s 
            \n\nSincerely,\nEMIF Catalogue
        """ % (name,request.user.get_full_name(), link_activation)
        # Send email to admins
        #send_mail(subject, message, settings.DEFAULT_FROM_EMAIL, emails_to_feedback)
        # Send email to user with the copy of feedback message
        send_mail(subject, message, settings.DEFAULT_FROM_EMAIL, [from_email])

    except BadHeaderError:
        return HttpResponse('Invalid header found.')



    return render(request, template_name, {'request': request, 'breadcrumb': True})

def sharedb_activation(request, activation_code, template_name="sharedb_invited.html"):

    if (request.user==None):
        return HttpResponse('You need to be authenticated.')
    __objs = SharePending.objects.filter(activation_code=activation_code, pending=True, user=request.user)
    if (len(__objs)==0):
        return HttpResponse('It is already activated or does the item has been expired.')

    if (len(__objs)>1):
        return HttpResponse('An error has occured. Contact the EMIF Catalogue Team.')
    
    sp = __objs[0]
    c = CoreEngine()
    results = c.search_fingerprint('id:' + sp.db_id)
    for r in results:
        _aux = r
        break
    
    _aux['user_t'] =  _aux['user_t'] + " \\ " + sp.user.username
    print _aux['user_t']

    c.update(r)
    sp.pending = False
    sp.save()


    try:
        subject = "EMIF Catalogue: Accepted database shared"
        message = """Dear %s,\n\n
            \n\n
            %s has been activated. You can access the new database in "Workspace" -> My Databases".
            \n\nSincerely,\nEMIF Catalogue
        """ % (request.user.get_full_name(), _aux['database_name_t'] )


        message_to_inviter = """Dear %s,\n\n
            \n\n
            %s has accepted to work with you in database %s. 
            
            \n\nSincerely,\nEMIF Catalogue
        """ % (sp.user_invite.get_full_name(), request.user.get_full_name(), _aux['database_name_t'])

        # Send email to admins
        send_mail(subject, message_to_inviter, settings.DEFAULT_FROM_EMAIL, [sp.user_invite.email])
        # Send email to user with the copy of feedback message
        send_mail(subject, message, settings.DEFAULT_FROM_EMAIL, [sp.user.email])

    except BadHeaderError:
        return HttpResponse('Invalid header found.')


    return render(request, template_name, {'request': request, 'breadcrumb': True})


# Documentation
def docs_api(request, template_name='docs/api.html'):
    return render(request, template_name, {'request': request, 'breadcrumb': True})


def clean_str_exp(s):
    return s.replace("\n", "|").replace(";", ",").replace("\t", "    ").replace("\r","").replace("^M","")

def save_answers_to_csv(list_databases, filename):
    """
    Method to export answers of a given database to a csv file
    """
    # Create the HttpResponse object with the appropriate CSV header.
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="EMIF_Catalogue_%s_%s.csv"' % (filename, datetime.datetime.now().strftime("%Y%m%d-%H%M%S"))

    if list_databases:
        writer = csv.writer(response, delimiter = '\t')
        writer.writerow(['DB_ID', 'DB_name', 'Questionset', 'Question', 'QuestioNumber', 'Answer'])
        for t in list_databases:
            id = t.id

            qsets, name, db_owners, fingerprint_ttype = createqsets(id, clean=False)

            for group in qsets.ordered_items():
                (k, qs) = group
                if (qs!=None and qs.list_ordered_tags!= None):
                    list_aux = sorted(qs.list_ordered_tags)
                    #import pdb
                    #pdb.set_trace()
                    for q in list_aux:
                        _answer = clean_str_exp(str(q.value))
                        if (_answer == "" and q.ttype=='comment'):
                            _answer = "-"
                        writer.writerow([id, name, k.replace('h1. ', ''), clean_str_exp(str(q.tag)), str(q.number), _answer])
            writer.writerow([id, name, "System", "Date", "99.0", t.date])
            writer.writerow([id, name, "System", "Date Modification", "99.1", t.date_modification])
            writer.writerow([id, name, "System", "Type", "99.2", t.type_name])
            writer.writerow([id, name, "System", "Type Identifier", "99.3", t.ttype])

    return response


def export_all_answers(request):
    """
    Method to export all databases answers to a csv file
    """

    list_databases = get_databases_from_solr(request, "*:*")
    return save_answers_to_csv(list_databases, "DBs")


def export_my_answers(request):
    """
    Method to export my databases answers to a csv file
    """

    user = request.user
    list_databases = get_databases_from_solr(request, "user_t:" + user.username)

    return save_answers_to_csv(list_databases, "MyDBs")


def export_bd_answers(request, runcode):
    """
    Method to export answers of a specific database to a csv file
    :param request:
    :param runcode:
    """

    list_databases = get_databases_from_solr(request, "id:" + runcode)
    return save_answers_to_csv(list_databases, 'MyDB')


class QuestionNumber:
    """
    State machine to create number of questions dynamically
    """
    def __init__(self):
        """
        n1, n2, n3, n4: level count
        t0, t1, t2, t3, t4: level text
        state: level state
        nQuestion: result
        :rtype : object
        """
        self._n1 = self._n2 = self._n3 = self._n4 = 1
        self._t0 = self._t1 = self._t2 = self._t3 = self._t4 = ''
        self._state = 'h1'
        self._nQuestion = ''
            
    def saveQuestionNumber(self):
        self._nQuestion = ''
        if self._t0:
            self._nQuestion += self._t0
        if self._t1 != '':
            self._nQuestion += '.' + self._t1
        if self._t2:
            self._nQuestion += '.' + self._t2
        if self._t3:
            self._nQuestion += '.' + self._t3
        if self._t4:
            self._nQuestion += '.' + self._t4

    def resetH0(self, hValue=1):
        self._t0 = str(hValue)
        self._n1 = self._n2 = self._n3 = self._n4 = 1
        self._t1 = self._t2 = self._t3 = self._t4 = ''
        self._state = 'h1'

    def resetH1(self):
        self._t1 = str(self._n1)
        self._n2 = self._n3 = self._n4 = 1
        self._t2 = self._t3 = self._t4 = ''
        self._n1 += 1
        self._state = 'h1'
        self.saveQuestionNumber()

    def resetH2(self):
        self._t2 = str(self._n2)
        self._n2 += 1
        self._n3 = self._n4 = 1
        self._t3 = self._t4 = ''
        self._state = 'h2'
        self.saveQuestionNumber()

    def resetH3(self):
        self._t3 = str(self._n3)
        self._n3 += 1
        self._n4 = 1
        self._t4 = ''
        self._state = 'h3'
        self.saveQuestionNumber()

    def resetH4(self):
        self._t4 = str(self._n4)
        self._n4 += 1
        self._state = 'h4'
        self.saveQuestionNumber()

    def getNumber(self, headingValue, hValue=1):
        """
        Function to get number of question, subquestion, etc.
        """
        # headingValue = 'h0' : QuestionSet
        if headingValue == 'h0':
            self.resetH0(hValue)
        elif self._state == 'h1':
            if headingValue == 'h1':
                self.resetH1()
            elif headingValue == 'h2':
                self.resetH2()
        elif self._state == 'h2':
            if headingValue == 'h1':
                self.resetH1()
            elif headingValue == 'h2':
                self.resetH2()
            elif headingValue == 'h3':
                self.resetH3()
        elif self._state == 'h3':
            if headingValue == 'h1':
                self.resetH1()
            elif headingValue == 'h2':
                self.resetH2()
            elif headingValue == 'h3':
                self.resetH3()
            elif headingValue == 'h4':
                self.resetH4()
        elif self._state == 'h4':
            if headingValue == 'h1':
                self.resetH1()
            elif headingValue == 'h2':
                self.resetH2()
            elif headingValue == 'h3':
                self.resetH3()
            elif headingValue == 'h4':
                self.resetH4()

        return self._nQuestion


def writeLog(log):
    with open("log_%s.txt" % datetime.datetime.now().strftime("%Y%m%d-%H%M%S"), "w") as f:
        f.write(log)
        f.close()

def get_slug(slug, q_id=None):
    slugs_objs = Slugs.objects.filter(slug1=slug, question__questionset__questionnaire=q_id)
    slug_aux = None
    if (len(slugs_objs)>0):
        slug_aux=slugs_objs[0]
    else:
        return slug
    slug_name_final = ""
    slug_arr = slug_aux.slug1.split("_")
    if len(slug_arr)>0:
        if (slug_arr[len(slug_arr)-1].isdigit()):
            slug_number = int(slug_arr[len(slug_arr)-1]) + 1
            slug_arr[len(slug_arr)-1] = str(slug_number)
            slug_name_final = "_".join(slug_arr)

        else:
            slug_name_final = slug_aux.slug1 + "_0"
    else:
        slug_name_final = slug_aux.slug1 + "_0"

    return slug_name_final

def save_slug(slugName, desc, question):
    slugsAux = Slugs()
    slugsAux.slug1 = slugName
    slugsAux.description = desc
    slugsAux.question = question
    slugsAux.save()


def import_questionnaire(request, template_name='import_questionnaire.html'):
    """
    To-Do
    - validation of template structure and content fields
    """

    #_debug=True to not save
    _debug = False

    qNumber = QuestionNumber()
    slugs = []
    from openpyxl import load_workbook
    from django.template.defaultfilters import slugify
    # wb = load_workbook(filename = r'/Volumes/EXT1/Dropbox/MAPi-Dropbox/EMIF/Code/emif/emif/questionnaire_ad_v2.xlsx')
    # wb = load_workbook(filename = r'/Volumes/EXT1/Dropbox/MAPi-Dropbox/EMIF/Observational_Data_Sources_Template_v5.xlsx')
    # wb = load_workbook(filename = r'C:/Questionnaire_template_v3.4.xlsx')
    wb = load_workbook(filename =r'/Volumes/EXT1/trash/Questionnaire_template_v3.5.3xlsx')
    ws = wb.get_active_sheet()
    log = ''

    # Cell B1: Name of questionnaire
    name = ws.cell('B1').value
    slugQ = convert_text_to_slug(ws.cell('B1').value)
    disable = False

    def format_number(number):
        # print number
        number_arr = number.split(".")
        
        result = number_arr[0] + "."
        for i in range(1,len(number_arr)):
            
            val = int(number_arr[i])
            if val<10:
                val = "0" + str(val)
            number_arr[i] = str(val)
            if (i!=len(number_arr)-1):
                result += str(val) + "."
            else:
                result += str(val)
        # print "result " + result
        return result

    questionnaire = Questionnaire(name=name, disable=disable, slug=slugQ, redirect_url='/')
    log += '\nQuestionnaire created %s ' % questionnaire

    try:
        if not _debug:
            questionnaire.save()
            log += '\nQuestionnaire saved %s ' % questionnaire
        _choices_array = {}
        _questions_rows = {}

        #############################
        # TIPS:
        # Type of Row: QuestionSet, Category, Question
        # Columns: Type, Text/Question, Level/Number, Data type, Value list, Help text/Description, Tooltip, Slug, Stats
        #############################

        for row in ws.rows[2:]:
            if len(row) > 0:
                type_Column = row[0]
                
                text_question_Column = row[1]

                if (text_question_Column.value!=None):
                    text_question_Column.value = text_question_Column.value.encode('ascii', 'ignore')
                level_number_column = row[2]
                _checks = ''

                # Type = QUESTIONSET
                # Columns required:  Type, Text/Question
                # Columns optional:  Help text/Description, Tooltip
                if str(type_Column.value) == "QuestionSet":
                    sortid = str(level_number_column.value)
                    try:
                        qNumber.getNumber('h0', sortid)
                    except:
                        writeLog(log)
                        raise
                    text_en = 'h1. %s' % text_question_Column.value
                    slug_qs = str(slugQ) + "_" + convert_text_to_slug(str(text_question_Column.value))
                    if row[5].value:
                        helpText = row[5].value
                    else:
                        helpText = ""
                    tooltip = False
                    if row[6].value:
                        if str(row[6].value).lower() == 'yes':
                            tooltip = True

                    questionset = QuestionSet(questionnaire=questionnaire, checks='required', sortid=sortid, text_en=text_en, heading=slug_qs, help_text=helpText, tooltip=tooltip)
                    log += '\n%s - QuestionSet created %s - %s ' % (type_Column.row, sortid, text_en)
                    try:
                        if not _debug:
                            questionset.save()
                            log += '\n%s - QuestionSet saved %s - %s ' % (type_Column.row, sortid, text_en)
                    except:

                        log += "\n%s - Error to save questionset %s - %s" % (type_Column.row, sortid, text_en)
                        writeLog(log)
                        raise

                # Type = CATEGORY
                # Columns required:  Type, Text/Question, Level/Number, Category
                # Columns optional:  Help text/Description, Slug, Tooltip, Dependencies
                elif str(type_Column.value) == "Category":

                    try:
                        text_en = str(level_number_column.value) + '. ' + str(text_question_Column.value)
                        if row[7].value:
                            slug = row[7].value
                        else:
                            slug = convert_text_to_slug(str(row[1].value)[:50])
                            slug = get_slug(slug, questionnaire.pk)

                        if row[5].value:
                            helpText = row[5].value
                        else:
                            helpText = ""
                        # print "HELP_TEXT1: " + str(helpText)
                        _tooltip = False

                        if row[6].value:
                            if str(row[6].value).lower() == 'yes':
                                _tooltip = True

                        #If has dependencies
                        if row[8].value:
                            try:
                                dependencies_list = row[8]
                                list_dep_aux = dependencies_list.value.split('|')
                                question_num_parent = str(_questions_rows.get(int(list_dep_aux[0])))

                                # print str(row[0].row) + " str(list_dep_aux[0]: " + str(int(list_dep_aux[0]))
                                # print str(row[0].row) + " question_num_parent: " + question_num_parent
                                # print str(row[0].row) + " _questions_rows: " + str(_questions_rows)

                                index_aux = int(str(list_dep_aux[1]))-1
                                choice_parent_list = _choices_array.get(int(list_dep_aux[0]))
                                choice_parent = choice_parent_list[index_aux]
                                _checks = 'dependent=\"' + str(question_num_parent) + ',' + str(choice_parent) + '\"'

                            except:
                                raise

                        try:
                            questionNumber = qNumber.getNumber(level_number_column.value)
                            questionNumber = format_number(str(questionNumber))
                        except:
                            log += "\n%s - Error to create Category number %s" % (type_Column.row, text_en)
                            writeLog(log)
                            raise
                        question = Question(questionset=questionset, text_en=text_en, number=str(questionNumber),
                                            type='comment', help_text=helpText, slug=slug, stats=False, category=True,
                                            tooltip=_tooltip, checks=_checks)

                        if not _debug:
                            question.save()
                            log += '\n%s - Category created %s ' % (type_Column.row, question)

                        _questions_rows[type_Column.row] = str(questionNumber)

                        if not _debug:
                            save_slug(question.slug,  question.text_en, question)

                        # slugs.append((question.slug,  question.text_en, question))
                        log += '\n%s - Category saved %s ' % (type_Column.row, question)
                    except:
                        log += "\n%s - Error to save Category %s" % (type_Column.row, text_en)
                        writeLog(log)
                        raise

                # Type = QUESTION
                # Columns required:  Type, Text/Question, Level/Number, Data Type, Category, Stats
                # Columns optional:  Value List, Help text/Description, Tooltip, Dependencies
                else:
                    #try:
                    text_en = str(level_number_column.value) + '. ' + str(text_question_Column.value)
                    #except:
                        #import pdb
                        #pdb.set_trace()
                    try:
                        dataType_column = row[3]
                        if row[7].value:
                            slug = row[7].value
                        else:
                            slug = convert_text_to_slug(str(row[1].value)[:50])
                            slug = get_slug(slug, questionnaire.pk)

                        if row[5].value:
                            helpText = row[5].value
                        else:
                            helpText = ''
                        
                        _tooltip = False

                        if row[6].value:
                            if str(row[6].value).lower() == 'yes':
                                _tooltip = True

                        #If has dependencies
                        if row[8].value:
                            try:
                                dependencies_list = row[8]
                                list_dep_aux = dependencies_list.value.split('|')
                                question_num_parent = str(_questions_rows.get(int(list_dep_aux[0])))

                                # print "###############"
                                # print str(row[0].row) + " dependencies_list: " + str(dependencies_list)
                                # print str(row[0].row) + " str(list_dep_aux[0]: " + str(int(list_dep_aux[0]))
                                # print str(row[0].row) + " question_num_parent: " + question_num_parent
                                # print str(row[0].row) + " _questions_rows: " + str(_questions_rows)
                                # print str(row[0].row) + " _choices_array: " + str(_choices_array)

                                index_aux = int(str(list_dep_aux[1]))-1
                                choice_parent_list = _choices_array.get(int(list_dep_aux[0]))
                                choice_parent = choice_parent_list[index_aux]
                                _checks = 'dependent=\"' + str(question_num_parent) + ',' + str(choice_parent) + '\"'
                            except:
                                raise

                        try:
                            questionNumber = qNumber.getNumber(level_number_column.value)
                            questionNumber = format_number(str(questionNumber))
                        except:
                            log += "\n%s - Error to create question number %s" % (type_Column.row, text_en)
                            writeLog(log)
                            raise

                        question = Question(questionset=questionset, text_en=text_en, number=str(questionNumber),
                                            type=dataType_column.value, help_text=helpText, slug=slug, stats=True,
                                            category=False, tooltip=_tooltip, checks=_checks)

                        log += '\n%s - Question created %s ' % (type_Column.row, question)

                        if not _debug:
                            question.save()

                        _questions_rows[type_Column.row] = str(questionNumber)

                        if not _debug:
                            save_slug(question.slug,  question.text_en, question)

                        # slugs.append((question.slug,  question.text_en, question))
                        log += '\n%s - Question saved %s ' % (type_Column.row, question)
                        if dataType_column.value in ['choice', 'choice-freeform', 'choice-multiple', 'choice-multiple-freeform']:
                            _choices_array_aux = []
                            # Parse of values list
                            values_list = row[4]
                            if (values_list!=None and values_list.value!=None):
                                list_aux = values_list.value.split('|')
                                i = 1
                                for ch in list_aux:
                                    try:
                                        choice = Choice(question=question, sortid=i, text_en=ch, value=ch)
                                        log += '\n%s - Choice created %s ' % (type_Column.row, choice)
                                        if not _debug:
                                            choice.save()
                                        _choices_array_aux.append(ch)

                                        log += '\n%s - Choice saved %s ' % (type_Column.row, choice)
                                        i += 1
                                    except:
                                        log += "\n%s - Error to save Choice %s" % (type_Column.row, choice)
                                        writeLog(log)
                                        raise
                                _choices_array[type_Column.row] = _choices_array_aux

                        if dataType_column.value in ['choice-yesno', 'choice-yesnocomment',
                                                             'choice-yesnodontknow']:
                            _choices_array[type_Column.row] = ['yes', 'no', 'dontknow']
                    except:
                        log += "\n%s - Error to save question %s" % (type_Column.row, text_en)
                        
                        writeLog(log)
                        raise

    except:
        log += '\nError to save questionsets and questions of the questionnaire %s ' % questionnaire
        writeLog(log)
        raise

    # for a in slugs:
    #     (_slug, _desc, question) = a
    #     # Write Slug
    #     slugsAux = Slugs()
    #     slugsAux.slug1 = _slug
    #     slugsAux.description = _desc
    #     slugsAux.question = question
    #     slugsAux.save()

    log += '\nQuestionnaire %s, questionsets, questions and choices created with success!! ' % questionnaire
    writeLog(log)
    # print log

    return render_to_response(template_name, {'import_questionnaire': True,
                              'request': request, 'breadcrumb': True}, RequestContext(request))


