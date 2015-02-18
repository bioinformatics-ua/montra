# -*- coding: utf-8 -*-
# Copyright (C) 2014 Universidade de Aveiro, DETI/IEETA, Bioinformatics Group - http://bioinformatics.ua.pt/
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with this program.  If not, see <http://www.gnu.org/licenses/>.
def import_questionnaire():

    from openpyxl import load_workbook
    wb = load_workbook(filename = 'EMIF-AD.xlsx')
    list_sheets = wb.get_sheet_names()

    ws = wb.get_active_sheet()
    print ws
    content = []


    for ws_name in list_sheets:
        ws = wb.get_sheet_by_name(ws_name)
        for row in ws.rows:
            for c in row:
                if c.value != None:
                    print c.value

        #content.append([c.value for c in row])

    # for row in ws.iter_rows(): # it brings a new method: iter_rows()
    #
    #     row_content = []
    #     for cell in row:
    #         row_content.append(cell.internal_value)
    #         print cell.row
    #     #     # content.append(cell.internal_value)
    #     content.append(cell.row)
    #     content[cell.row].append(row_content)
    #print content


import_questionnaire()
