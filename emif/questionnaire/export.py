# -*- coding: utf-8 -*-
# Copyright (C) 2014 Luís A. Bastião Silva and Universidade de Aveiro
#
# Authors: Luís A. Bastião Silva    <bastiao@ua.pt>
#          Ricardo Ribeiro          <ribeiro.r@ua.pt>
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with this program.  If not, see <http://www.gnu.org/licenses/>.
#

from questionnaire.models import *

import os
import re

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Style, PatternFill, Alignment, Font, Border, Side
from openpyxl.cell import Cell
from openpyxl.worksheet.datavalidation import DataValidation, ValidationType

# TODO: put this code in utils
def removehs(value):
    value = value.replace('h0. ','')
    value = value.replace('h1. ','')
    value = value.replace('h2. ','')
    value = value.replace('h3. ','')
    value = value.replace('h4. ','')
    value = value.replace('h5. ','')
    value = value.replace('h6. ','')
    value = value.replace('h7. ','')
    return value


"""This class is used to export the fingerprint template to excel
"""
class ExportQuestionnaire(object):


    def __init__(self, questionnaire, file_path):
        self.questionnaire = questionnaire
        self.file_path = file_path

    def ignore_questionsets(self, questionset_list):
        raise NotImplementedError("Please Implement this method")

    def export(self):
        raise NotImplementedError("Please Implement this method")


    def level(self, question):
        return question.number.count(".") + 1


    def clean(self, text):
        return removehs(text)


    """This method will build the object according with the type
    of the object to export.
    """
    @staticmethod
    def factory(t_type, questionnaire, file_path):
        if t_type == "csv_plain":
            return ExportQuestionnaireCSVPlain(questionnaire, file_path)

        elif t_type == 'excel':
            return ExportQuestionnaireExcel(questionnaire, file_path)
        else:
            raise Exception("The supplied format is not supported")


class ExportQuestionnaireCSVPlain(ExportQuestionnaire):

    def __init__(self, questionnaire, file_path):
        ExportQuestionnaire.__init__(self, questionnaire)

    def ignore_questionsets(self, questionset_list):
        raise NotImplementedError("Please Implement this method")

    def export(self):
        result = ""
        questionsets = self.questionnaire.questionsets()

        for qs in questionsets:

            result += str(qs.sortid) + " - " + self.clean(qs.text)  + "\n"
            questions = qs.questions()
            for q in questions:
                _level  = self.level(q)
                result += self.get_tabs(_level)
                result += str(q.number)  + " " + self.clean(q.text) +"\n"


        f = open(self.file_path, 'w')
        f.write(result)
        f.close()


class ExportQuestionnaireExcel(ExportQuestionnaire):

    # dependency questions need to be able to translate question numbers into excel line numbers.
    __defaultstyle = Style(font=Font(name='Verdana', size=8),
            alignment=Alignment(wrap_text=True),
            border=Border(
                left=Side(border_style='thin', color='FF000000'),
                right=Side(border_style='thin', color='FF000000'),
                top=Side(border_style='thin', color='FF000000'),
                bottom=Side(border_style='thin', color='FF000000')
                )
            )
    __boldstyle = Style(font=Font(bold=True))

    __headerstyle = Style(alignment=Alignment(horizontal='center'),
            fill=PatternFill(fill_type='solid', start_color='FFCCCCCC'))

    __validatetype = DataValidation(type="list", formula1='"QuestionSet, Category, Question"', allow_blank=True)
    __validateyesno = DataValidation(type="list", formula1='"Yes, No"', allow_blank=True)


    __validateqtype = DataValidation(type="list", formula1='"open, open-button, open-upload-image, open-textfield, choice-yesno, choice-yesnocomment, choice-yesnodontknow, comment, choice, choice-freeform, choice-multiple, choice-multiple-freeform, range, timeperiod, publication, sameas, custom, datepicker"', allow_blank=True)


    def __init__(self, questionnaire, file_path):
        ExportQuestionnaire.__init__(self, questionnaire, file_path)
        self.__ignore_questionsets = []

        self.__number_map = {}

    def __boolean_to_string(self, value):
        if value == True:
            return 'yes'
        elif value == False:
            return 'no'

        return 'error'

    def __setDefaultStyle(self, _cell):
        _cell.style = self.__defaultstyle

    def __setBold(self, _cell):
        _cell.style = self.__defaultstyle.copy(font=Font(name='Verdana', size=8, bold=True))

    def __setHeader(self, _cell):
        _cell.style.font.bold = True

        # Cell background color
        _cell.style = self.__headerstyle

    def __setColumnSizes(self, ws, sizes):
        columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']

        for i in xrange(len(columns)):
            ws.column_dimensions[columns[i]].width = sizes[i]
    def __getChoices(self, question):
        retstring = ""

        choices = Choice.objects.filter(question=question)

        for choice in choices:
            retstring+=choice.value+'|'

        if(len(retstring) > 1):
            return retstring[:-1]

        return retstring

    def __getChoiceNumber(self, parent, option):
        yesno_questions = ['choice-yesno','choice-yesnocomment','choice-yesnodontknow']

        if parent.type in yesno_questions:
            if option.lower() == 'yes':
                return "1"
            elif option.lower() == 'no':
                return "2"
            elif option.lower() == 'dontknow':
                return "3"

            return "error"
        else:
            return Choice.objects.get(question=parent, value=option).sortid


    def __processDependencies(self, question):

        try:
            if (question.checks == None):
                question.checks = ""
            valid = re.search('.*dependent="([0-9\.]+),(.*)".*', question.checks, re.IGNORECASE)

            if valid:
                number = valid.group(1)
                option = valid.group(2)

                (line, parent) = self.__number_map[number]

                optionid = self.__getChoiceNumber(parent, option)

                restring = str(line)+'|'+str(optionid)

                return restring

            # No dependencies
            else:
                return ""
        except KeyError:
            print "-- ERROR: Couldn't find a mapping for question "+str(question.number)
            return "error"

    def __addQuestion(self, line, ws, question):
        self.__number_map[question.number] = (line, question)

        valid = re.search('(h[0-9])+\. (.*)', question.text_en, re.IGNORECASE)
        choice_types = ['choice', 'choice-freeform', 'choice-multiple', 'choice-multiple-freeform', 'choice-multiple-freeform-options']

        if valid:
            level = str(valid.group(1))
            text = str(valid.group(2))
            type = None

            choices = ''

            if question.category:
                type = 'Category'
            else:
                type = 'Question'

            if question.type in choice_types:
                choices = self.__getChoices(question)

            ws.append([
                    type,
                    text,
                    level,
                    question.type,
                    choices,
                    question.help_text,
                    self.__boolean_to_string(question.tooltip),
                    question.slug_fk.slug1,
                    self.__processDependencies(question),
                    '',
                    '',
                ])

            for row in ws.iter_rows('A'+str(line)+":k"+str(line)):
                for cell in row:
                    self.__setDefaultStyle(cell)

            if question.category:
                self.__setBold(ws.cell('B'+str(line)))

            return True

        return False

    def ignore_questionsets(self, questionset_list):
        self.__ignore_questionsets = questionset_list

    def export(self):
        wb = load_workbook(filename =r'questionnaire/empty2.xlsx')
        ws = wb.get_active_sheet()
        ws.title = "Questionnaire"

        ws.cell('B1').value = self.questionnaire.name
        self.__setBold(ws.cell('B1'))

        # for sanity, im keeping a pointer to the row im in...
        pointer = 3

        for questionset in self.questionnaire.questionsets():
            if questionset in self.__ignore_questionsets:
                continue

            ws.append(['QuestionSet', questionset.text_en.replace('h1. ',''),
                        questionset.sortid, '', '', questionset.help_text.replace('<br />', '\n'),
                        self.__boolean_to_string(questionset.tooltip), '',
                        '', '', '' ])

            for row in ws.iter_rows('A'+str(pointer)+":k"+str(pointer)):
                for cell in row:
                    self.__setDefaultStyle(cell)

            self.__setBold(ws.cell('A'+str(pointer)))
            self.__setBold(ws.cell('B'+str(pointer)))
            self.__setBold(ws.cell('C'+str(pointer)))

            pointer += 1

            for question in questionset.questions():
                inserted = self.__addQuestion(pointer, ws, question)

                if inserted:
                    pointer += 1
                else:
                    print "-- ERROR PROCESSING QUESTION header for: "+str(question.text_en)
                    break

        # Adding validation data, to create dropdown abilities as the original
        self.__validatetype.ranges.append('A3:A'+str(pointer))
        self.__validateqtype.ranges.append('D3:D'+str(pointer))
        self.__validateyesno.ranges.append('G3:G'+str(pointer))

        ws.add_data_validation(self.__validatetype)
        ws.add_data_validation(self.__validateyesno)
        ws.add_data_validation(self.__validateqtype)

        # Freezing first two rows
        ws.freeze_panes = ws.cell('A3')

        wb.save(self.file_path)



# def main():

#     q = Questionnaire.objects.get(id=53)

#     exporter = ExportQuestionnaire.factory("csv_plain", q, '/tmp/csvplain.csv')
#     exporter.export()









